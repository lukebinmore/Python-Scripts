# region Imports
import os
import sys
import json
from datetime import date, datetime
import msvcrt

LIBS = os.path.join(os.path.dirname(__file__), os.path.join("AppData", "Required Libraries"))
sys.path.insert(0, LIBS)
from openpyxl import load_workbook

# endregion


# region CLASSES
class Student:
    def __init__(self, firstname, lastname, id, group):
        self.firstname = firstname
        self.lastname = lastname
        self.id = id
        self.group = group
        self.badges = []
        self.accountFound = True
        self.outstanding = 0
        self.late = 0

    def __str__(self):
        return f"{self.lastname}, {self.firstname} ({self.id} / {self.group})"

    def to_dict(self):
        return {"id": self.id, "badges": [b.to_dict() for b in self.badges]}


class Badge:
    def __init__(self, name, date):
        self.name = name
        try:
            self.date = datetime.strptime(date, "%d-%m-%Y").date().isoformat()
        except ValueError:
            self.date = datetime.strptime(date, "%Y-%m-%d").date().isoformat()

    def __str__(self):
        d = date.fromisoformat(self.date).strftime("%d-%m-%Y")
        return f"{self.name} (Date: {d})"

    def to_dict(self):
        return {"name": self.name, "date": self.date}

    def __eq__(self, other):
        return isinstance(other, Badge) and self.name == other.name

    def __hash__(self):
        return hash(self.name)


class Group:
    def __init__(self, group):
        self.groupName = group
        self.students = []

    def __str__(self):
        return f"{self.groupName} - {len(self.students)} Students"

    def AddStudent(self, student):
        self.students.append(student)

    def SortStudents(self):
        self.students.sort(key=lambda student: (student.lastname, student.firstname))


class MenuArgs:
    def __init__(self, page=0):
        self.page = page
        self.totalPages = 0
        self.goBack = False
        self.optionResult = None


class Filters:
    def __init__(self):
        self.Clear()

    def Clear(self):
        self.Groups = []
        self.Badges = []
        self.Statuses = []
        self.Students = []


# endregion


# region Globals
CURRENTDIR = os.path.dirname(os.path.abspath(__file__))
SETTINGSPATH = os.path.join(CURRENTDIR, "AppData/Settings.json")

settings = {}
schedule = []
students = []
groups = []
statuses = [("Completed", "G"), ("Late", "Y"), ("Outstanding", "R"), ("Missing Accounts", "M")]
latestHM = None
filters = Filters()
# endregion


# region SETUP FUNCTIONS
def ImportSettings():
    global settings
    print("Importing Settings...", end="\r")

    ThrowError("Settings file not found.") if not os.path.exists(SETTINGSPATH) else None

    with open(SETTINGSPATH, "r") as settingsFile:
        settings = json.load(settingsFile)

    print("Importing Settings... Done")


def ImportSchedule():
    global schedule
    print("Importing Schedule...", end="\r")

    schedulePath = os.path.join(CURRENTDIR, settings["schedule_path"])

    if not os.path.exists(schedulePath):
        with open(schedulePath, "w") as tmpFile:
            json.dump([], tmpFile)
            tmpFile.close()

    with open(schedulePath, "r") as scheduleFile:
        scheduleRaw = json.load(scheduleFile)

        for badgeData in scheduleRaw:
            schedule.append(Badge(badgeData["name"], badgeData["date"]))

    SortSchedule()

    print("Importing Schedule... Done")


def ImportStudents():
    global students, groups
    print("Importing Student Data [")

    studentDataPath = os.path.join(CURRENTDIR, settings["student_data_path"])
    dataFilesFound = False

    if not os.path.exists(studentDataPath):
        os.makedirs(studentDataPath)

    for filename in os.listdir(studentDataPath):
        if filename.endswith(".xlsx"):
            print(f" - Importing {filename}...", end="\r")

            dataFilesFound = True
            filePath = os.path.join(studentDataPath, filename)
            groupName = filename.replace(".xlsx", "")

            group = Group(groupName)
            groups.append(group)

            workbook = load_workbook(filePath, True, True)
            worksheet = workbook.active

            headers = GetHeaderIndicies(worksheet, settings["student_data_headers"])

            for row in worksheet.iter_rows(min_row=2):
                firstname = row[headers["firstname"]].value
                lastname = row[headers["lastname"]].value
                id = row[headers["id"]].value
                student = Student(firstname, lastname, id, groupName)
                students.append(student)
                group.AddStudent(student)

            print(f" - Importing {filename}... Done")

    if not dataFilesFound:
        ThrowError("No student data files found.")

    print("]... Done")


def ImportDataHistory():
    global students
    print("Importing Data History...", end="\r")

    dataHistoryPath = os.path.join(CURRENTDIR, settings["data_history_path"])
    if not os.path.exists(dataHistoryPath):
        return

    with open(dataHistoryPath, "r") as file:
        fileData = json.load(file)
        for studentData in fileData:
            student = next((s for s in students if s.id == studentData["id"]), None)
            if not student:
                continue

            for badgeData in studentData["badges"]:
                badge = Badge(badgeData["name"], badgeData["date"])
                student.badges.append(badge)

    print("Importing Data History... Done")


def ImportAnalytics():
    global students
    print("Importing Analytics Data [")

    dataFileFound = False

    for filename in os.listdir(CURRENTDIR):
        if filename.endswith(".xlsx"):
            print(f" - Importing {filename}...", end="\r")

            dataFileFound = True
            missing = 0
            filePath = os.path.join(CURRENTDIR, filename)
            workbook = load_workbook(filePath, True, True)
            worksheet = workbook.active

            headers = GetHeaderIndicies(worksheet, settings["analytics_data_headers"])

            for student in students:
                rowData = None
                for row in worksheet.iter_rows(min_row=2):
                    if student.id == row[headers["email"]].value.split("@")[0]:
                        rowData = row

                if not rowData:
                    student.accountFound = False
                    missing += 1
                    continue

                badges = rowData[headers["badges"]].value
                badges = badges.split(",") if badges else []

                for badge in badges:
                    badge = Badge(badge, date.today().isoformat())
                    if badge not in student.badges:
                        student.badges.append(badge)

            print(f" - Importing {filename}... Done - {missing} Missing Accounts")

    if not dataFileFound:
        ThrowError("No analytics data files found.")

    print("]... Done")


def CalculateBadgeData():
    global students
    print("Calculating Outstanding Badge Counts...")

    for student in students:
        print(f" - Checking {student}...", end="\r")
        for reqBadge in schedule:
            if reqBadge not in student.badges:
                student.outstanding += 1
            else:
                badgeIdx = student.badges.index(reqBadge)
                stuBadge = student.badges[badgeIdx]

                if stuBadge.date > reqBadge.date:
                    student.late += 1
        print(f" - Checking {student}... Done")

    print("Calculating Outstanding Badge Counts... Done")


def UpdateDataHistory():
    print("Updating Data History...", end="\r")
    filePath = os.path.join(CURRENTDIR, settings["data_history_path"])

    with open(filePath, "w", encoding="utf-8") as file:
        json.dump(
            [s.to_dict() for s in students],
            file,
            indent=2,
            ensure_ascii=False,
        )

    print("Updating Data History... Done")


def GetLatestHomework():
    global latestHW
    print("Getting Latest Homework...", end="\r")

    latestHW = schedule[-1] if len(schedule) > 0 else None
    now = date.today().isoformat()
    for badge in schedule:
        if badge.date > latestHW.date and badge.date <= now:
            latestHW = badge

    print("Getting Latest Homework... Done")


# endregion


# region Utility Functions
def ThrowError(message):
    print(f"\n\nERROR: {message}\n\n")
    print("Please fix the issue and try again.")
    print("Press any key to exit...")
    Getch()
    exit()


def Clear():
    os.system("cls" if os.name == "nt" else "clear")


def ClearLines(n=1):
    for _ in range(n):
        sys.stdout.write("\x1b[1A")
        sys.stdout.write("\x1b[2K")
    sys.stdout.flush()


def Getch():
    key = msvcrt.getch()
    if key == b"\xe0" or key == b"\x00":
        key2 = msvcrt.getch()
        if key2 == b"K":
            return "LEFT"
        elif key2 == b"M":
            return "RIGHT"
        else:
            return f"EXT_{ord(key2)}"
    elif key == b"\x1b":
        return "ESC"
    elif key == b"\r":
        return "1"
    elif key == b"\x08":
        return "BACKSPACE"
    else:
        return key.decode("utf-8").upper()


def Colour(text, colour):
    if not colour:
        return text

    c = str(colour).upper()
    if c == "G":
        code = "32"
    elif c == "R":
        code = "31"
    elif c == "Y":
        code = "33"
    elif c == "M":
        code = "95"
    elif c == "H":
        return f"\033[97;44m{text}\033[0m"

    return f"\033[{code}m{text}\033[0m"


def GetHeaderIndicies(ws, headers):
    headerRow = next(ws.iter_rows(min_row=1, max_row=1))
    headerIndicies = {cell.value: idx for idx, cell in enumerate(headerRow)}
    results = {}

    for header in headers:
        try:
            results[header] = headerIndicies[headers[header]]
        except:
            ThrowError(f"Header '{headers[header]}' not found in worksheet.")

    return results


def SortSchedule():
    global schedule
    schedule.sort(key=lambda b: b.date, reverse=True)


def UpdateSchedule():
    print("\nUpdating Schedule...", end="\r")
    filePath = os.path.join(CURRENTDIR, settings["schedule_path"])

    with open(filePath, "w", encoding="utf-8") as file:
        json.dump(
            schedule,
            file,
            indent=2,
            ensure_ascii=False,
            default=lambda o: o.__dict__,
        )

    print("Updating Schedule... Done")


def PressAnyKey(value=""):
    print()
    print(f"Invalid input ({value}), Please try again.") if value else ""
    print("Press any key to continue...")
    Getch()

    lines = 4 if value else 3
    ClearLines(lines)


def ConfirmPrompt(prompt=""):
    while True:
        if prompt:
            print(f"{prompt} (Y/N)")
        else:
            print("Confirm? (Y/N)")

        ch = Getch()
        if ch == "Y":
            return True
        elif ch == "N":
            return False
        else:
            PressAnyKey(ch)


def SwapGrouping():
    if settings["grouping"] == "Group":
        settings["grouping"] = "Homework"
    else:
        settings["grouping"] = "Group"

    with open(SETTINGSPATH, "w", encoding="utf-8") as settingsFile:
        json.dump(
            settings,
            settingsFile,
            indent=2,
            ensure_ascii=False,
        )


# endregion


# region MENU BUILDING FUNCTIONS
def MHeader(title):
    Clear()
    print(Colour("#### IDEA AWARD HOMEWORK TRACKER V2.0 ####", "H"))
    print(f"## {title} ##\n")


def MOptions(args, options):
    args.totalPages = (len(options) + 8) // 9

    start = args.page * 9
    pageOptions = options[start : start + 9]

    for i, option in enumerate(pageOptions, 1):
        if len(option) >= 3:
            colour = "G" if option[2] else "R"
            print(f"{i} - {Colour(option[0], colour)}")
        else:
            print(f"{i} - {option[0]}")

    pageOpt = "\n"
    pageOpt += "PREV(←) - " if args.page > 0 else ""
    pageOpt += f"Page: {args.page + 1}"
    pageOpt += " - NEXT(→)" if args.page < args.totalPages - 1 else ""
    print(pageOpt) if args.totalPages > 1 else None

    print("\nESC - Exit Program")
    print("BACKSPACE - Go Back")

    return args


def MInput(args, options):
    print("\nPlease select an option: ", end="")
    ch = Getch()

    args.goBack = False

    if ch == "ESC":
        exit()
    elif ch == "BACKSPACE":
        args.goBack = True
    elif ch == "LEFT" and args.page > 0:
        args.page = (args.page - 1) % args.totalPages
    elif ch == "RIGHT" and args.page < args.totalPages - 1:
        args.page = (args.page + 1) % args.totalPages
    elif ch.isdigit():
        idx = int(ch)
        if (len(options) - args.page * 9) >= idx:
            args.optionResult = options[args.page * 9 + idx - 1][1]()
            if args.optionResult and args.optionResult == "goBack":
                args.goBack = True

    return args


def ShowFilters():
    output = ""
    for filterType in sorted(f for f in dir(filters) if not f.startswith("_")):
        values = getattr(filters, filterType)
        if not callable(values) and len(values) > 0:
            if len(values) < 5:
                output += f" - {filterType}:\n"
                for filter in values:
                    output += f"    - {filter}\n"
                output += "\n"
            else:
                output += f" - {filterType}: {len(values)} filters applied\n"

    if output != "":
        print("Negative Filters Applied:")
        print(output)


def PrintStudent(student, badge):
    stdBadge = next((b for b in student.badges if b == badge), None)
    status = statuses[2]
    if stdBadge:
        if stdBadge.date <= badge.date:
            status = statuses[0]
        else:
            status = statuses[1]

    status = statuses[3] if not student.accountFound else status

    if status[0] not in filters.Statuses:
        output = f" - {Colour(student, status[1])}"
        if status != statuses[3]:
            if student.late > 0:
                output += " - " + Colour(f"{student.late} Late", "Y")
            if student.outstanding > 0:
                output += " - " + Colour(f"{student.outstanding} Outstanding", "R")
        else:
            output += Colour(" - Account Not Found", "M")

        print(output)


def ViewHomework():
    MHeader("Homework Results")
    ShowFilters()

    reqGroups = [g for g in groups if g not in filters.Groups]
    reqBadges = [b for b in schedule if b not in filters.Badges]

    if settings["grouping"] == "Homework":
        for badge in reqBadges:
            print(f"=== Homework: {badge} ===")
            for group in reqGroups:
                print(f"--- Group: {group} ---")
                for student in group.students:
                    if student not in filters.Students:
                        PrintStudent(student, badge)
                print()
            print()
    else:
        for group in reqGroups:
            print(f"=== Group: {group} ===")
            for badge in reqBadges:
                print(f"--- Homework: {badge} ---")
                for student in group.students:
                    if student not in filters.Students:
                        PrintStudent(student, badge)
                print()
            print()

    PressAnyKey()


def ViewLatestHomework():
    MHeader(f"Latest Homework: {latestHW.name if latestHW else 'N/A'}")

    for group in groups:
        print(f"--- Group: {group} ---")
        for student in group.students:
            PrintStudent(student, latestHW)
        print()

    PressAnyKey()


def AddScheduleBadge():
    global schedule

    name = ""
    date = ""

    while True:
        MHeader("Add Homework")
        name = input("Enter the badge name (or leave blank to cancel): ").strip()
        if not name:
            return

        if not ConfirmPrompt():
            continue
        else:
            break

    while True:
        MHeader("Add Homework")
        print(f"Badge Name: {name}\n")
        date = input("Enter the due date (DD-MM-YYYY) (or leave blank to cancel): ").strip()
        if not date:
            return

        try:
            date = datetime.strptime(date, "%d-%m-%Y").date().isoformat()
        except ValueError:
            PressAnyKey(date)
            continue

        if not ConfirmPrompt():
            continue
        else:
            break

    schedule.append(Badge(name, date))
    SortSchedule()
    UpdateSchedule()


def EditScheduleBadgeName(badgeIdx):
    global schedule
    while True:
        MHeader("Edit Badge Name")
        print(f"Current Name: {schedule[badgeIdx].name}\n")
        newName = input("Enter new badge name (or leave blank to cancel): ").strip()
        if not newName:
            return
        if ConfirmPrompt():
            schedule[badgeIdx].name = newName
            UpdateSchedule()
            break


def EditScheduleBadgeDueDate(badgeIdx):
    global schedule
    while True:
        MHeader("Edit Badge Due Date")
        currentDate = date.fromisoformat(schedule[badgeIdx].date).strftime("%d-%m-%Y")
        print(f"Current Due Date: {currentDate}\n")
        newDate = input("Enter new due date (DD-MM-YYYY) (or leave blank to cancel): ").strip()
        if not newDate:
            return
        try:
            newDate = datetime.strptime(newDate, "%d-%m-%Y").date().isoformat()
        except ValueError:
            PressAnyKey(newDate)
            continue
        if ConfirmPrompt():
            schedule[badgeIdx].date = newDate
            badge = schedule[badgeIdx]
            SortSchedule()
            UpdateSchedule()
            return schedule.index(badge)


def DeleteScheduleBadge(badgeIdx):
    global schedule
    MHeader("Delete Badge")
    if ConfirmPrompt(f"Are you sure you want to delete the badge '{schedule[badgeIdx].name}'?"):
        del schedule[badgeIdx]
        UpdateSchedule()

    return "goBack"


def ToggleFilter(filterType, value):
    filterList = getattr(filters, filterType)
    if isinstance(value, list):
        filterList[:] = value
    elif value in filterList:
        filterList.remove(value)
    else:
        filterList.append(value)


# endregion


# region MENU'S
def ViewHomeworkMenu(title):
    args = MenuArgs()
    while True:
        newGrouping = "Homework" if settings["grouping"] == "Group" else "Group"
        options = [
            ["View Homework Results", ViewHomework],
            ["Filters", FilterTypesMenu],
            [f"Organise by {newGrouping}", SwapGrouping],
            ["Clear Filters", filters.Clear],
        ]

        MHeader(title)
        print(f"Current Organising Setting: {settings['grouping']}\n")
        args = MOptions(args, options)
        args = MInput(args, options)

        if args.goBack:
            break


def FilterMenu(filterType):
    args = MenuArgs()
    while True:
        filterList = getattr(filters, filterType)

        sourceList = []
        if filterType == "Groups":
            sourceList = groups
        elif filterType == "Badges":
            sourceList = schedule
        elif filterType == "Statuses":
            sourceList = [s[0] for s in statuses]
        elif filterType == "Students":
            sourceList = students

        allStatus = len(filterList) == 0
        options = [["Filter All", lambda: ToggleFilter(filterType, sourceList if allStatus else []), allStatus]]
        for val in sourceList:
            if filterType == "Students":
                groupNames = [g.groupName for g in filters.Groups]
                if val.group in groupNames:
                    continue

            options.append([val, lambda v=val: ToggleFilter(filterType, v), val not in filterList])

        MHeader(f"Filter by {filterType}")
        args = MOptions(args, options)
        args = MInput(args, options)

        if args.goBack:
            break


def FilterTypesMenu():
    args = MenuArgs()
    while True:
        options = [
            ["Filter by Groups", lambda: FilterMenu("Groups")],
            ["Filter by Homework", lambda: FilterMenu("Badges")],
            ["Filter by Status", lambda: FilterMenu("Statuses")],
            ["Filter by Students", lambda: FilterMenu("Students")],
        ]

        MHeader("Filters Menu")
        args = MOptions(args, options)
        args = MInput(args, options)

        if args.goBack:
            break


def EditSheduleBadgeMenu(badgeIdx):
    args = MenuArgs()
    while True:
        options = [
            ["Edit Name", lambda: EditScheduleBadgeName(badgeIdx)],
            ["Edit Due Date", lambda: EditScheduleBadgeDueDate(badgeIdx)],
            [
                Colour("Delete Badge", "R"),
                lambda: DeleteScheduleBadge(badgeIdx),
            ],
        ]

        MHeader("Edit Homework")
        print(f"Editing Badge: {schedule[badgeIdx].name}\n")
        args = MOptions(args, options)
        args = MInput(args, options)

        if args.optionResult != None:
            badgeIdx = args.optionResult

        if args.goBack:
            break


def ScheduleMenu(title):
    args = MenuArgs()
    while True:
        options = [["Add Homework", AddScheduleBadge]]

        for i, badge in enumerate(schedule):
            options.append([badge.__str__(), lambda i=i: EditSheduleBadgeMenu(i)])

        MHeader(title)
        args = MOptions(args, options)
        args = MInput(args, options)

        if args.goBack:
            break


# endregion


# region WRAPPERS
def Setup():
    ImportSettings()
    ImportSchedule()
    ImportStudents()
    ImportDataHistory()
    ImportAnalytics()
    CalculateBadgeData()
    UpdateDataHistory()
    GetLatestHomework()


def MainMenu():
    args = MenuArgs()
    while True:
        options = [
            ["View Latest Homework", ViewLatestHomework],
            ["View Homework", lambda: ViewHomeworkMenu("View Homework Results")],
            ["Schedule Manager", lambda: ScheduleMenu("Schedule Manager")],
        ]

        MHeader("Main Menu")
        args = MOptions(args, options)
        args = MInput(args, options)


# endregion

# region MAIN
Setup()
MainMenu()
# endregion
