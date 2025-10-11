# region Library Imports
import os
import sys

current_dir = os.path.dirname(os.path.abspath(__file__))
libs_path = os.path.join(current_dir, "Required Libraries")
sys.path.insert(0, libs_path)
print(sys.path)

import openpyxl

# endregion

# region Global Variables
CURRENT_CODE = "Crispin25"
STUDENT_DATA_HEADERS = ["Last Name", "First Name", "Candidate Number"]
ANALYTICS_DATA_HEADERS = ["Code", "Email", "Badge List"]
OUTPUT_FILENAME = "Results.txt"
print_history = "iDea Award Progress Checker"
students = []
analytics = []
groups = []
STUDENT_DATA_FOLDER = os.path.join(current_dir, "Students")
# endregion


# region Classes
class Student:
    def __init__(self, firstname, lastname, id_number, group):
        self.firstname = firstname
        self.lastname = lastname
        self.id_number = id_number
        self.group = group
        self.badges = []

    def __str__(self):
        return f"{self.firstname} {self.lastname} ({self.id_number}) - {self.group}"


class Group:
    def __init__(self, group):
        self.group_name = group
        self.students = []

    def AddStudent(self, student):
        self.students.append(student)


# endregion


# region Functions
def Update_Print_History(message):
    global print_history
    print_history += message


def Print_Main_Screen():
    global print_history
    os.system("cls" if os.name == "nt" else "clear")
    print(print_history)


def GetRequiredBadges():
    badges = []
    Update_Print_History("\n\nCurrent required badges:")

    while True:
        Print_Main_Screen()

        badge = (
            input(
                "\nPlease enter a required badge (or leave blank to finish): "
            )
            .strip()
            .lower()
        )
        if badge != "":
            badges.append(badge)
            Update_Print_History(f"\n - {badge}")
        else:
            if len(badges) == 0:
                input(
                    "\nYou must enter at least one badge. (Press Enter to continue)"
                )
            else:
                return badges


def LoadStudents():
    global students, groups
    Update_Print_History("\n\nCollecting student data:")

    students = []

    for file in os.listdir(STUDENT_DATA_FOLDER):
        Print_Main_Screen()

        if file.endswith(".xlsx") or file.endswith(".xlsm"):
            group_name = file[0:-5]
            groups.append(Group(group_name))
            print(f" - Loading student data for {file}...")
            filepath = os.path.join(STUDENT_DATA_FOLDER, file)

            workbook = openpyxl.load_workbook(filepath, data_only=True)
            worksheet = workbook.active

            header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
            header_indices = {
                cell.value: idx for idx, cell in enumerate(header_row)
            }

            for row in worksheet.iter_rows(min_row=2):
                values = [
                    row[header_indices[header]].value.lower().strip()
                    for header in STUDENT_DATA_HEADERS
                ]

                lastname, firstname, id_number = values
                students.append(
                    Student(firstname, lastname, id_number, group_name)
                )

            Update_Print_History(
                f"\n - Loading student data for {file} - Done!"
            )


def LoadAnalytics():
    global analytics
    Update_Print_History("\n\nCollecting analytics data:")

    analytics = []
    global current_dir

    for file in os.listdir(current_dir):
        Print_Main_Screen()
        if file.endswith(".xlsx") or file.endswith(".xlsm"):
            print(f" - Loading analytics data...")
            filepath = os.path.join(current_dir, file)

            workbook = openpyxl.load_workbook(filepath, data_only=True)
            worksheet = workbook.active

            header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
            header_indices = {
                cell.value: idx for idx, cell in enumerate(header_row)
            }

            for row in worksheet.iter_rows(min_row=2):
                values = [
                    row[header_indices[header]].value
                    for header in ANALYTICS_DATA_HEADERS
                ]

                values[1] = values[1].split("@")[0].lower().strip()
                if values[2] == None:
                    values[2] = []
                else:
                    values[2] = values[2].lower().split(",")

                if values[0] == CURRENT_CODE:
                    analytics.append(values)

            Update_Print_History(
                f"\n - Loading analytics data for {file} - Done!"
            )


def MatchStudentData():
    global students, analytics
    Update_Print_History("\n\nMatching student data with analytics data:")
    Print_Main_Screen()

    for student in students:
        print(f" - Matching data for {student}...", end="\r")
        for entry in analytics:
            if student.id_number in entry[1]:
                student.badges = entry[2]
        print(f" - Matching data for {student} - Done!")

    Update_Print_History("\n - Matching data - Done!")


def CheckProgress(required_badges):
    global students
    Update_Print_History("\n\nChecking student progress:")
    Print_Main_Screen()

    for student in students:
        print(f" - Checking progress for {student}...", end="\r")

        for badge in required_badges:
            if badge not in student.badges:
                break
        else:
            students.remove(student)

        print(f" - Checking progress for {student} - Done!")

    Update_Print_History("\n - Checking progress - Done!")


def SortGroups():
    global groups, students
    Update_Print_History("\n\nSorting students into groups:")
    Print_Main_Screen()

    for group in groups:
        print(f" - Sorting students into {group.group_name}...", end="\r")

        for student in students:
            if student.group == group.group_name:
                group.AddStudent(student)
        print(f" - Sorting students into {group.group_name} - Done!")

    Update_Print_History("\n - Sorting students - Done!")


def GenerateOutput(required_badges):
    global groups

    Update_Print_History("\n\nGenerating output data:")
    Print_Main_Screen()

    output_string = "Badges required for completion:"

    for badge in required_badges:
        output_string += f"\n - {badge}"

    output_string += "\n\nStudents with outstanding badges:"

    for group in groups:
        print(f" - Generating output data for {group.group_name}...", end="\r")
        output_string += f"\n\nGroup: {group.group_name}"
        if len(group.students) == 0:
            output_string += (
                "\n - All students have completed the required badges!"
            )
        else:
            for student in group.students:
                output_string += f"\n - {student}"
            output_string += "\n"
        print(f" - Generating output data for {group.group_name} - Done!")

    Update_Print_History("\n - Generating output data - Done!")

    return output_string


def ExportResults(output_string):
    global current_dir
    Update_Print_History("\n\nExporting results to file:")
    Print_Main_Screen()

    output_filepath = os.path.join(current_dir, OUTPUT_FILENAME)
    with open(output_filepath, "w") as file:
        file.write(output_string)

    Update_Print_History(f"\n - Exporting results - Done!")
    Update_Print_History(f"\n - File saved to: {output_filepath}")

    Print_Main_Screen()

    print("\nResults:\n\n" + output_string)


def Main():
    global students, analytics
    required_badges = GetRequiredBadges()
    LoadStudents()
    LoadAnalytics()
    MatchStudentData()
    CheckProgress(required_badges)
    SortGroups()
    output_string = GenerateOutput(required_badges)
    ExportResults(output_string)
    input()


# endregion

# region Main Code

Main()

# endregion
